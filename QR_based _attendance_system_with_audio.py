import PySimpleGUI as sg
from pyzbar.pyzbar import decode
import cv2
import sounddevice as sd
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import os
import pyttsx3

def load_attendance_list(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')
    return df.iloc[:, 0].tolist()

def mark_attendance(file_path, name, present_file, present_names):
    if name in present_names:
        return present_names

    if os.path.exists(present_file):
        present_workbook = load_workbook(present_file)
    else:
        present_workbook = Workbook()

    present_sheet = present_workbook.active

    # Find the next empty cell in column A
    row_num = 1
    while present_sheet.cell(row=row_num, column=1).value is not None:
        row_num += 1

    # Write the name in the next empty cell
    present_sheet.cell(row=row_num, column=1, value=name)
    present_workbook.save(present_file)

    present_names.add(name)
    return present_names

def beep():
    duration = 0.1
    frequency = 1000
    sample_rate = 44100
    t = np.linspace(0, duration, int(sample_rate * duration), False)
    beep_signal = 0.7 * np.sin(2 * np.pi * frequency * t)
    sd.play(beep_signal, sample_rate)
    sd.wait()

def get_camera_list():
    camera_list = []
    for i in range(10):
        cap = cv2.VideoCapture(i)
        if cap.isOpened():
            camera_list.append(f"Camera {i}")
            cap.release()
    return camera_list

def create_present_window(present_names):
    sg.theme("DarkGrey5")

    layout = [
        [sg.Text("Present List", size=(40, 1), justification="center", font=("Helvetica", 16))],
        [sg.Listbox(present_names, size=(40, 15), key="-PRESENT_LIST-")]
    ]

    return sg.Window("Present List", layout, finalize=True)

def update_present_window(window, present_names):
    window["-PRESENT_LIST-"].update(present_names)

def display_attendance_table(attendance_list, present_names):
    attendance_status = []

    for name in attendance_list:
        status = "Checked In" if name in present_names else "Not Checked In"
        attendance_status.append([name, status])

    sg.theme("DarkGrey5")

    layout = [
        [sg.Text("Attendance List Comparison", size=(40, 1), justification="center", font=("Helvetica", 16))],
        [sg.Table(values=attendance_status,
                  headings=["Name", "Status"],
                  auto_size_columns=False,
                  justification="right",
                  num_rows=min(25, len(attendance_list)) if attendance_list else 1)]
    ]

    window = sg.Window("Attendance List Comparison", layout, finalize=True)

    while True:
        event, values = window.read()

        if event in (sg.WIN_CLOSED, "Close"):
            break

    window.close()

def scan_qr_code(attendance_list, file_path, present_file, selected_camera="Camera 0"):
    sg.theme("DarkGrey5")

    layout = [
        [sg.Image(filename="", key="-IMAGE-")],
        [sg.Text("Scan a QR Code", size=(40, 1), justification="center", font=("Helvetica", 16))],
        [sg.Button("OK", size=(20, 1)), sg.Button("Exit", size=(20, 1)),
         sg.Text("Select Camera:"), sg.Combo(get_camera_list(), default_value=selected_camera, key="-CAMERA-")]
    ]

    window = sg.Window("QR Code Scanner", layout, finalize=True)

    cap = cv2.VideoCapture(0)
    present_names = []
    present_window = create_present_window(present_names)
    last_attendance_list = attendance_list[:]

    # Initialize the text-to-speech engine
    tts_engine = pyttsx3.init()

    while True:
        event, values = window.read(timeout=20)

        if event in (sg.WIN_CLOSED, "Exit"):
            break

        selected_camera = values["-CAMERA-"]
        if selected_camera != selected_camera:
            cap.release()
            cap = cv2.VideoCapture(int(selected_camera.split()[-1]))

        ret, frame = cap.read()

        if not ret:
            print("Error capturing video frame.")
            break

        decoded_objects = decode(frame)
        if decoded_objects:
            data = decoded_objects[0].data.decode('utf-8').strip()

            if data in attendance_list:
                if data not in present_names:
                    present_names.append(data)
                    mark_attendance(file_path, data, present_file, present_names)
                    welcome_message = f"WELCOME, {data}! Thanks for coming!"
                    # Speak out the welcome message
                    tts_engine.say(welcome_message)
                    tts_engine.runAndWait()
                    beep()
                    update_present_window(present_window, present_names)

                sg.popup(f"Scanned: {data}\n{welcome_message}", title="Scan Result")
            else:
                not_welcome_message = f"You are not welcome, please scan again."
                # Speak out the not welcome message and "Scan again"
                tts_engine.say("Scan again")
                tts_engine.runAndWait()
                beep()
                sg.popup(f"Scanned: {data}\n{not_welcome_message}", title="Scan Result")

        if event == "OK":
            display_attendance_table(attendance_list, present_names)

        if last_attendance_list != attendance_list:
            last_attendance_list = attendance_list[:]

        if last_attendance_list == present_names and last_attendance_list != []:
            sg.popup("Let's start the celebration!")

        if frame is not None:
            imgbytes = cv2.imencode(".png", frame)[1].tobytes()
            window["-IMAGE-"].update(data=imgbytes)

    cap.release()
    window.close()

    # Save the present attendees to a local file
    with open("Present_Attendees.txt", "w") as present_file:
        for name in present_names:
            present_file.write(name + "\n")

if __name__ == "__main__":
    input_file = input("Enter the path to the Excel file with the attendance list: ")
    present_file = "Present_Attendees.txt"  # Specify the file to save present attendees
    attendance_list = load_attendance_list(input_file)
    scan_qr_code(attendance_list, input_file, present_file)  # No need to pass selected_camera as an argument
