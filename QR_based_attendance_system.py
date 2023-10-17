import PySimpleGUI as sg
from pyzbar.pyzbar import decode
import cv2
import sounddevice as sd
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import os

# Load the attendance list from an Excel file
def load_attendance_list(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')
    return set(df.iloc[:, 0].tolist())

def mark_attendance(file_path, name, present_file, present_names):
    if name in present_names:
        return present_names

    if os.path.exists(present_file):
        present_workbook = load_workbook(present_file)
    else:
        present_workbook = Workbook()

    present_sheet = present_workbook.active

    # Find the next empty cell in column A
    row_num = 1
    while present_sheet.cell(row=row_num, column=1).value is not None:
        row_num += 1

    # Write the name in the next empty cell
    present_sheet.cell(row=row_num, column=1, value=name)
    present_workbook.save(present_file)

    present_names.add(name)
    return present_names

def check_already_checked_in(name, present_names):
    return name in present_names

def beep():
    # Play a louder beep sound
    duration = 0.1  # seconds
    frequency = 1000  # Hz
    sample_rate = 44100
    t = np.linspace(0, duration, int(sample_rate * duration), False)
    beep_signal = 0.7 * np.sin(2 * np.pi * frequency * t)  # Increase amplitude (0.7)
    sd.play(beep_signal, sample_rate)
    sd.wait()

def scan_qr_code(attendance_list, file_path, present_file):
    sg.theme("DarkGrey5")

    layout = [
        [sg.Image(filename="", key="-IMAGE-")],
        [sg.Text("Scan a QR Code", size=(40, 1), justification="center")],
        [sg.Text("", size=(40, 1), justification="center", key="-MESSAGE-", font=("Helvetica", 16))],
        [sg.Button("OK", size=(20, 1)), sg.Button("Exit", size=(20, 1))]
    ]

    window = sg.Window("QR Code Scanner", layout, finalize=True)

    cap = cv2.VideoCapture(0)
    message_element = window["-MESSAGE-"]

    present_file_path = None
    present_names = set()

    while True:
        event, values = window.read(timeout=20)

        if event in (sg.WIN_CLOSED, "Exit"):
            break

        ret, frame = cap.read()

        if not ret:
            print("Error capturing video frame.")
            break

        decoded_objects = decode(frame)
        if decoded_objects:
            data = decoded_objects[0].data.decode('utf-8').strip()
            if check_already_checked_in(data, present_names):
                message = "You have already checked in"
                beep()  # Play a louder beep sound
            else:
                present_names = mark_attendance(file_path, data, present_file, present_names)
                if data in attendance_list:
                    message = f"WELCOME, {data}\nThanks for coming!"
                    beep()  # Play a louder beep sound
                else:
                    message = "NOT WELCOME"

            sg.popup(f"Scanned: {data}\n{message}", title="Scan Result")
            message_element.update("Scan a QR Code")

        if event == "OK":
            message_element.update("Scan a QR Code")

        if frame is not None:
            imgbytes = cv2.imencode(".png", frame)[1].tobytes()
            window["-IMAGE-"].update(data=imgbytes)

    cap.release()
    window.close()

if __name__ == "__main__":
    input_file = input("Enter the path to the Excel file with the attendance list: ")
    attendance_list = load_attendance_list(input_file)
    present_file = "Present_Attendees.xlsx"
    scan_qr_code(attendance_list, input_file, present_file)
